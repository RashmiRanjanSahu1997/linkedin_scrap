import time
import os
import random
import requests
import zipfile
import pandas as pd
import pickle
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from selenium.common.exceptions import NoAlertPresentException, NoSuchElementException, TimeoutException, ElementNotInteractableException


class LinkedInAutomation:
    def __init__(self, email, password, data_directory, cookie_file='cookies.pkl'):
        self.email = email
        self.password = password
        self.data_directory = data_directory
        self.cookie_file = cookie_file
        self.driver = self._initialize_driver()

    def _initialize_driver(self):
        chrome_options = self._configure_chrome_options()
        return webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)

    def _configure_chrome_options(self):
        chrome_options = Options()
        prefs = {
            "download.default_directory": self.data_directory,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
        }
        chrome_options.add_experimental_option("prefs", prefs)
        return chrome_options

    def login(self, use_cookies=False):
        self.driver.get('https://www.linkedin.com/login')
        if use_cookies:
            try:
                cookies = self._load_cookies()
                for cookie in cookies:
                    self.driver.add_cookie(cookie)
                print("Cookies loaded. Attempting to login with cookies...")
                self.driver.refresh()
                self._wait_for_login_complete()
                return
            except Exception as e:
                print("Failed to load cookies:", e)

        email_input = self.driver.find_element(By.ID, 'username')
        password_input = self.driver.find_element(By.ID, 'password')
        email_input.send_keys(self.email)
        password_input.send_keys(self.password)
        password_input.send_keys(Keys.RETURN)
        print("Please enter the 2FA code manually on LinkedIn.")
        time.sleep(15)
        self._wait_for_login_complete()
        self._save_cookies()

    def _save_cookies(self):
        with open(self.cookie_file, 'wb') as file:
            pickle.dump(self.driver.get_cookies(), file)
        print("Cookies saved successfully.")

    def _load_cookies(self):
        if os.path.isfile(self.cookie_file):
            with open(self.cookie_file, 'rb') as file:
                return pickle.load(file)
        return []

    def _wait_for_login_complete(self):
        try:
            WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.ID, 'feed-nav')))
            print("Login successful.")
        except Exception as e:
            print("Login failed or timed out:", e)
            print("Current URL:", self.driver.current_url)
            print("Page source:", self.driver.page_source[:1000])
            try:
                WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.global-nav__me-photo')))
                print("Login successful (fallback).")
            except Exception as e:
                print("Fallback also failed:", e)

    def download_data(self):
        self.driver.get('https://www.linkedin.com/mypreferences/d/download-my-data')
        try:
            iframes = self.driver.find_elements(By.TAG_NAME, 'iframe')
            print(f"Found {len(iframes)} iframes. Trying to switch...")
            for iframe in iframes:
                self.driver.switch_to.frame(iframe)
                try:
                    download_button = self.driver.find_element(By.CSS_SELECTOR, 'button.download-btn')
                    self.driver.execute_script("arguments[0].scrollIntoView(true);", download_button)
                    WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'button.download-btn')))
                    download_button.click()
                    print("Download button clicked within iframe.")
                    time.sleep(30)
                    latest_zip = self._get_latest_file(".zip")
                    self._extract_zip(latest_zip)
                    break
                except Exception as iframe_exception:
                    print(f"Failed to initiate download within iframe: {iframe_exception}")
                finally:
                    self.driver.switch_to.default_content()
        except Exception as e:
            print(f"Error handling iframes: {e}")

    def _extract_zip(self, zip_file):
        with zipfile.ZipFile(zip_file, 'r') as zip_ref:
            zip_ref.extractall(self.data_directory)
        print(f"Extracted {zip_file} to {self.data_directory}")

    def load_data(self):
        csv_files = [os.path.join(self.data_directory, f) for f in os.listdir(self.data_directory) if f.endswith('.csv')]
        data_frames = {}
        for csv_file in csv_files:
            try:
                with open(csv_file, 'r', encoding='utf-8') as f:
                    header_found = False
                    header_line = None
                    lines = []
                    for line_num, line in enumerate(f):
                        if not header_found:
                            if os.path.basename(csv_file) == 'Connections.csv' and line_num >= 3:
                                header_found = True
                                header_line = line.strip()
                            elif line.strip().startswith(("First Name", "Name", "Title", "Company", "CONVERSATION ID", 'From', 'Date', 'Organization', 'Endorsement Date')):
                                header_found = True
                                header_line = line.strip()
                        else:
                            lines.append(line.strip())
                    if header_line:
                        data = "\n".join([header_line] + lines)
                        from io import StringIO
                        df = pd.read_csv(StringIO(data))
                        data_frames[os.path.basename(csv_file)] = df
                    else:
                        print(f"Header not found in {csv_file}")
            except Exception as e:
                print(f"Error reading {csv_file}: {e}")
        return data_frames

    def _get_latest_file(self, file_extension):
        files = [os.path.join(self.data_directory, f) for f in os.listdir(self.data_directory) if f.endswith(file_extension)]
        if not files:
            raise FileNotFoundError(f"No files with extension {file_extension} found in directory {self.data_directory}")
        latest_file = max(files, key=os.path.getctime)
        print(f"Latest file found: {latest_file}")
        return latest_file

    def perform_actions(self, data_frames,csv_file):
        actions = []
        if csv_file == "Shares.csv":
            if 'Shares.csv' in data_frames:
                connections = data_frames['Shares.csv']
                messages = ["Hello! Thanks for the Endorsement!"]
                for index, row in connections.iterrows():
                    share_url = row['ShareLink']
                    messages = ["Hello! Thanks for Comments!"]
                    self.driver.get(share_url)
                    time.sleep(random.uniform(5, 10))
                    try:
                        comment_box = WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.ql-editor')))
                        random_message = random.choice(messages)
                        comment_box.send_keys(random_message)
                        print('Comment typed.')
                        send_button = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'artdeco-button') and contains(@class, 'artdeco-button--1') and contains(@class, 'artdeco-button--tertiary') and .//span[text()='Comment']]")))
                        self.driver.execute_script("arguments[0].scrollIntoView(true);", send_button)
                        time.sleep(1)
                        self.driver.execute_script("arguments[0].click();", send_button)
                        print('Comment button clicked.')
                    except NoAlertPresentException:
                        print('No alert present. Continuing...')
                    except Exception as e:
                        print(f'An error occurred: {e}')
                    time.sleep(random.uniform(2, 5))
            else:
                print('Not available shares file')

        elif csv_file == "Endorsement_Received_Info.csv":
            if 'Endorsement_Received_Info.csv' in data_frames:
                connections = data_frames['Endorsement_Received_Info.csv']
                messages = ["Hello! Thanks for the Endorsement! "]
                for index, row in connections.iterrows():
                    profile_url = row['Endorser Public Url']
                    msg = row['Skill Name']
                    messages = [f"Hello! Thanks for the Endorsement! {msg}"]
                    self.driver.get("https://" + profile_url)
                    time.sleep(random.uniform(5, 10))
                    try:
                        message_button = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'button.artdeco-button--primary.pvs-profile-actions__action')))
                        print('Message button found, attempting to click...')
                        message_button.click()
                        time.sleep(random.uniform(2, 5))
                        message_box = WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.msg-form__contenteditable')))
                        random_message = random.choice(messages)
                        message_box.send_keys(random_message)
                        send_button = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'button.msg-form__send-button.artdeco-button--1')))
                        send_button.click()
                        action = f"Visited {profile_url} and sent a message: {random_message}"
                        actions.append(action)
                    except Exception as e:
                        print(f'An error occurred while sending message: {e}')
                        continue
                    time.sleep(random.uniform(2, 5))
            else:
                print('not available this file')
        elif csv_file == 'Connections.csv':
            
            if 'Connections.csv' in data_frames:

                connections = data_frames['Connections.csv']
                messages = ["Hello! Nice to connect with you.", "Hi there! Hope you're doing well.", "Greetings! Looking forward to our connection.", "Hello! How are you?"]
                for index, row in connections.iterrows():
                    profile_url = row['URL']
                    self.driver.get(profile_url)
                    time.sleep(random.uniform(5, 10))
                    try:
                        print('Attempting to find the message button...')
                        message_button = WebDriverWait(self.driver, 20).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, 'button.artdeco-button--primary.pvs-profile-actions__action'))
                        )
                        print('Message button found, attempting to click...')
                        message_button.click()
                        time.sleep(random.uniform(2, 5))
                        message_box = WebDriverWait(self.driver, 30).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, 'div.msg-form__contenteditable'))
                        )
                        random_message = random.choice(messages)
                        message_box.send_keys(random_message)
                        send_button = WebDriverWait(self.driver, 20).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, 'button.msg-form__send-button.artdeco-button--1'))
                        )
                        send_button.click()    
                    except Exception as e:
                        print(f"Failed to send message to {profile_url}: {e}")
                    time.sleep(random.uniform(2, 5))
            else:
                print(' not available')
        elif csv_file == "messages.csv":
            if 'messages.csv' in data_frames:
                endorsements = data_frames['messages.csv']
                processed_conversations =[]
                for index, row in endorsements.iterrows():
                    conversation_id = row['CONVERSATION ID']
                    if conversation_id not in processed_conversations:
                        processed_conversations.append(conversation_id)
                        self.driver.get(f'https://www.linkedin.com/messaging/thread/{conversation_id}')
                        time.sleep(random.uniform(5, 10))

                        try:
                            message_box = WebDriverWait(self.driver, 20).until(
                                EC.presence_of_element_located((By.CSS_SELECTOR, 'div.msg-form__contenteditable'))
                            )
                            message_box.send_keys("Thank you for your message. I will get back to you shortly.")
                            
                            send_button = WebDriverWait(self.driver, 20).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, 'button.msg-form__send-button.artdeco-button--1')))
                            send_button.click()
                        except Exception as e:
                            print(f"Failed to send message to conversation {conversation_id}: {e}")
                        time.sleep(random.uniform(2, 5))
        elif csv_file =='Invitations.csv':
            if 'Invitations.csv' in data_frames:
                invitations = data_frames['Invitations.csv']
                for index, row in invitations.iterrows():
                    profile_url = row['inviteeProfileUrl']
                    self.driver.get(profile_url)
                    time.sleep(random.uniform(5, 10))
                    try:
                        # Check if the "Pending" button is present
                        pending_buttons = self.driver.find_elements(By.XPATH, "//button[contains(@aria-label, 'click to withdraw invitation')]")
                        pending_buttons = self.driver.find_elements(By.XPATH, "//button[contains(@aria-label, 'Pending, click to withdraw invitation')]")
                        if pending_buttons:
                            pending_button = pending_buttons[0]
                            print(f"Found pending button for {profile_url}: {pending_button}")
                            WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable(pending_button))
                            pending_button.click()
                            time.sleep(random.uniform(2, 5))
                            print("Pending button clicked on profile:", profile_url)

                            # Locate and click the "Withdraw" button
                            withdraw_button = WebDriverWait(self.driver, 20).until(
                                EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Withdraw')]"))
                            )
                            self.driver.execute_script("arguments[0].scrollIntoView(true);", withdraw_button)
                            withdraw_button.click()
                            actions.append(f"Withdrew invitation for {profile_url}")

                            # Debugging output
                            print("Withdraw button clicked on profile:", profile_url)
                        else:
                            actions.append(f"No pending button found for {profile_url}")

                    except (NoSuchElementException, TimeoutException, ElementNotInteractableException) as e:
                        actions.append(f"Failed to withdraw invitation for {profile_url}: {e}")
                        # Debugging output
                        print(f"Error on profile {profile_url}: {e}")
                    time.sleep(random.uniform(2, 5))
        
        self._save_actions_to_excel(actions)
        print('Actions performed:', actions)
        return actions

    def _save_actions_to_excel(self, actions):
        excel_file = os.path.join(self.data_directory, 'actions.xlsx')
        if os.path.exists(excel_file):
            wb = load_workbook(excel_file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
        for action in actions:
            ws.append([action])
        wb.save(excel_file)
        print(f'Actions saved to {excel_file}')


if __name__ == "__main__":
    EMAIL = "rashmiranjansahu1997@gmail.com"
    PASSWORD = "rashmi8249"
    DATA_DIRECTORY = "/home/ranjan/linkedin_scrap/data"
    automation = LinkedInAutomation(EMAIL, PASSWORD, DATA_DIRECTORY)
    if automation._load_cookies():

        automation.login(use_cookies=True)
    else:
        automation.login()
    # automation.download_data()
    data_frames = automation.load_data()
    csv_files = 'Invitations.csv'
    automation.perform_actions(data_frames,csv_file=csv_files)
